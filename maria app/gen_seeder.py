import openpyxl

wb = openpyxl.load_workbook('Backup_BBDD_TimeGuard_2026-04-17 (1).xlsx')

workers = []
ws = wb['BBDD_TRABAJADORES']
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 4)]
    if row[0]:
        workers.append(row)

entries = []
ws = wb['BBDD_JORNADAS']
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 9)]
    if row[0]:
        entries.append(row)

comps = []
ws = wb['BBDD_COMPENSACIONES']
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 7)]
    if row[0]:
        comps.append(row)

lines = []
lines.append('<?php')
lines.append('')
lines.append('namespace Database\\Seeders;')
lines.append('')
lines.append('use Illuminate\\Database\\Seeder;')
lines.append('use Illuminate\\Support\\Facades\\DB;')
lines.append('')
lines.append('class TimeguardBackupSeeder extends Seeder')
lines.append('{')
lines.append('    public function run(): void')
lines.append('    {')
lines.append('        // Disable foreign key checks for truncate')
lines.append("        DB::statement('SET FOREIGN_KEY_CHECKS=0');")
lines.append('')
lines.append('        // ---- TRABAJADORES ----')
lines.append("        DB::table('timeguard_workers')->truncate();")
lines.append("        DB::table('timeguard_workers')->insert([")
for w in workers:
    is_active = 1 if str(w[2]).upper() == 'TRUE' else 0
    name = str(w[1]).replace("'", "\\'")
    lines.append("            ['id' => '{}', 'name' => '{}', 'is_active' => {}, 'created_at' => now(), 'updated_at' => now()],".format(w[0], name, is_active))
lines.append('        ]);')
lines.append('')
lines.append('        // ---- JORNADAS (time entries) ----')
lines.append("        DB::table('timeguard_time_entries')->truncate();")
lines.append('        $entries = [')
for e in entries:
    lunch_manual = 1 if str(e[5]).upper() == 'TRUE' else 0
    free_day = 1 if str(e[6]).upper() == 'TRUE' else 0
    notes = 'null' if e[7] is None else "'" + str(e[7]).replace("'", "\\'") + "'"
    lines.append("            ['id' => '{}', 'worker_id' => '{}', 'date' => '{}', 'hours_brutas' => {}, 'lunch_discount' => {}, 'is_lunch_manual' => {}, 'is_free_day' => {}, 'notes' => {}, 'created_at' => now(), 'updated_at' => now()],".format(
        e[0], e[1], e[2], int(e[3]), int(e[4]), lunch_manual, free_day, notes))
lines.append('        ];')
lines.append('        foreach (array_chunk($entries, 100) as $chunk) {')
lines.append("            DB::table('timeguard_time_entries')->insert($chunk);")
lines.append('        }')
lines.append('')
lines.append('        // ---- COMPENSACIONES ----')
lines.append("        DB::table('timeguard_compensations')->truncate();")
lines.append("        DB::table('timeguard_compensations')->insert([")
for c in comps:
    notes = 'null' if c[5] is None else "'" + str(c[5]).replace("'", "\\'") + "'"
    lines.append("            ['id' => '{}', 'worker_id' => '{}', 'date' => '{}', 'type' => '{}', 'minutes' => {}, 'notes' => {}, 'created_at' => now(), 'updated_at' => now()],".format(
        c[0], c[1], c[2], c[3], int(c[4]), notes))
lines.append('        ]);')
lines.append('')
lines.append("        DB::statement('SET FOREIGN_KEY_CHECKS=1');")
lines.append('    }')
lines.append('}')

output_path = '../database/seeders/TimeguardBackupSeeder.php'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('Seeder generado: {} workers, {} entries, {} comps'.format(len(workers), len(entries), len(comps)))


